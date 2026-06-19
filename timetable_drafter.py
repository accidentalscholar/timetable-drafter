# ==========================================
# Timetable Drafter 
# Version: 1.2.0
# Citation: Pundir, V. (2026, June 19). Timetable Drafter Version (1.2.0). Retrieved from https://github.com/accidentalscholar/timetable-drafter.
# Citation: RIS and BibTeX files included for referencing software.
# Tested in: Python 3.12.11 64 bit packaged by Anaconda, Inc.
# Reporsitory: https://github.com/accidentalscholar/timetable-drafter
# Provided under: GNU AFFERO GENERAL PUBLIC LICENSE (see accompanying license file)
# ==========================================

import sys
import os
import site
import subprocess
import math
import warnings
import traceback
from functools import reduce
from datetime import datetime
import tkinter as tk
from tkinter import filedialog

# Suppress harmless openpyxl warnings globally
warnings.filterwarnings('ignore', category=UserWarning)

# ==========================================
# 0. FAULT TOLERANCE & PATH RESOLUTION
# ==========================================
user_site = site.getusersitepackages()
if user_site and os.path.exists(user_site) and user_site not in sys.path:
    sys.path.append(user_site) 
    
user_scripts = os.path.join(os.path.dirname(user_site), 'Scripts') if user_site else ""
if user_scripts and os.path.exists(user_scripts):
    os.environ["PATH"] += os.pathsep + user_scripts

path_file = os.path.join(os.getcwd(), 'path.txt')
if os.path.exists(path_file):
    print(f"[*] Found optional 'path.txt'. Loading custom paths...", flush=True)
    with open(path_file, 'r') as f:
        for line in f:
            custom_path = line.strip()
            if custom_path and not custom_path.startswith('#'):
                if os.path.exists(custom_path):
                    if custom_path not in sys.path:
                        sys.path.insert(0, custom_path)
                    os.environ["PATH"] += os.pathsep + custom_path
                    print(f"    -> Dynamically injected: {custom_path}", flush=True)

def install_and_import(package_name):
    try:
        __import__(package_name)
    except ImportError:
        print(f"[*] Missing '{package_name}'. Attempting auto-install...", flush=True)
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", package_name])
            print(f"[+] Successfully installed {package_name}!\n", flush=True)
            if user_site and os.path.exists(user_site) and user_site not in sys.path:
                sys.path.append(user_site)
        except Exception as e:
            sys.exit(f"\n[!] ERROR: Could not install '{package_name}'. Run '!pip install {package_name}' in your console.")

for lib in ["pandas", "pulp", "openpyxl"]:
    install_and_import(lib)

import pandas as pd
from pulp import LpProblem, LpMinimize, LpVariable, lpSum, PULP_CBC_CMD
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

def compute_lcm(x, y):
    return (x * y) // math.gcd(x, y)

def compute_lcm_list(lst):
    if not lst: return 1
    return reduce(compute_lcm, lst, 1)

# Helper function to brutally sanitize all Excel inputs to prevent hidden whitespace & NaN bugs
def sanitize_df(df):
    if df is not None:
        for col in df.select_dtypes(['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            # Destroy the phantom "nan" string created by Pandas when casting empty cells
            df[col] = df[col].replace({'nan': '', 'None': '', 'null': ''})
    return df

# ==========================================
# 1. GUI FILE SELECTOR
# ==========================================
def get_file_path():
    root = tk.Tk()
    root.withdraw()
    root.call('wm', 'attributes', '.', '-topmost', True)
    print("[*] Waiting for you to select the Input Excel file...", flush=True)
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    root.destroy() 
    if not file_path:
        sys.exit("[!] No file selected. Exiting.")
    return file_path

file_path = get_file_path()

# ==========================================
# 2. PRE-FLIGHT DATA VALIDATOR & SANITIZER
# ==========================================
print("[*] Performing Pre-Flight Data Validation...", flush=True)
try:
    xls = pd.ExcelFile(file_path)
except Exception as e:
    sys.exit(f"[!] FATAL: Could not open file. Details: {e}")

validation_errors = []
required_schema = {
    "Programmes": ["ProgrammeCode", "ProgrammeName"],
    "Staff": ["Initials", "FullName"],
    "Rooms": ["RoomType", "Capacity"],
    "Modules": ["ModuleCode", "ModuleName", "RoomType"],
    "Timeslots": ["Day", "TimeSlot"],
    "Module_Staff": ["ModuleCode", "StaffInitials", "AllocationType"],
    "Module_Programmes": ["ModuleCode", "ProgrammeCode", "Term"]
}

for sheet, required_cols in required_schema.items():
    if sheet not in xls.sheet_names:
        validation_errors.append(f"MISSING SHEET: '{sheet}' is required but not found in the workbook.")
    else:
        df_check = pd.read_excel(xls, sheet_name=sheet)
        missing_cols = [col for col in required_cols if col not in df_check.columns]
        if missing_cols:
            validation_errors.append(f"MISSING COLUMNS: In sheet '{sheet}', could not find {missing_cols}.")

if validation_errors:
    print("\n" + "="*50)
    print(" 🚨 PRE-FLIGHT CHECK FAILED: STRUCTURAL ERRORS")
    sys.exit(1)

# Ingest and aggressively sanitize all strings to prevent hidden Space/Tab blocking bugs
df_prog = sanitize_df(pd.read_excel(xls, sheet_name="Programmes").dropna(subset=["ProgrammeCode"]))
df_staff = sanitize_df(pd.read_excel(xls, sheet_name="Staff").dropna(subset=["Initials"]))
df_rooms = sanitize_df(pd.read_excel(xls, sheet_name="Rooms").dropna(subset=["RoomType"]))
df_modules = sanitize_df(pd.read_excel(xls, sheet_name="Modules").dropna(subset=["ModuleCode"]))
df_mod_staff = sanitize_df(pd.read_excel(xls, sheet_name="Module_Staff").dropna(subset=["ModuleCode", "StaffInitials"]))
df_mod_prog = sanitize_df(pd.read_excel(xls, sheet_name="Module_Programmes").dropna(subset=["ModuleCode", "ProgrammeCode"]))
df_timeslots = sanitize_df(pd.read_excel(xls, sheet_name="Timeslots"))

term_columns = [col for col in df_prog.columns if str(col).endswith('Cohort')]
if not term_columns:
    print("\n[!] FORMAT ERROR (Programmes): Missing 'Cohort' column.")
    sys.exit(1)

ordered_terms = [col.replace('Cohort', '') for col in term_columns]

print("[+] Pre-Flight Check Passed! Data is clean.", flush=True)

# ==========================================
# 3. DATA INGESTION & NORMALIZATION
# ==========================================
df_prog.set_index("ProgrammeCode", inplace=True)
df_staff.set_index("Initials", inplace=True)
df_rooms.set_index("RoomType", inplace=True)
df_modules.set_index("ModuleCode", inplace=True)

# Robust Staff Availability Parsing (with Term logic)
try:
    df_staff_avail = pd.read_excel(xls, sheet_name="Staff_Availability")
    df_staff_avail = sanitize_df(df_staff_avail)
    df_staff_avail.dropna(subset=['StaffInitials'], inplace=True)
    
    restricted_staff_set = set(df_staff_avail['StaffInitials'])
    permitted_staff = set()
    
    if 'Term' not in df_staff_avail.columns:
        df_staff_avail['Term'] = ''
        
    for _, row in df_staff_avail.iterrows():
        st = row['StaffInitials']
        t_str = f"{row['Day']}_{row['TimeSlot']}"
        term_val = str(row.get('Term', '')).strip()
        
        if not term_val or term_val.lower() in ['all', 'none']:
            for tm in ordered_terms:
                permitted_staff.add((st, tm, t_str))
        else:
            permitted_staff.add((st, term_val, t_str))
except Exception as e:
    permitted_staff = set()
    restricted_staff_set = set()
    
# Programme Availability Parsing
try:
    df_prog_avail = sanitize_df(pd.read_excel(xls, sheet_name="Programme_Availability").dropna(how='all'))
    permitted_progs = set(zip(df_prog_avail['ProgrammeCode'], df_prog_avail['Day']))
    restricted_prog_set = set(df_prog_avail['ProgrammeCode'])
except:
    permitted_progs = set()
    restricted_prog_set = set()

try:
    df_offset = sanitize_df(pd.read_excel(xls, sheet_name="Module_Cohort_Offset").dropna(how='all'))
    offset_dict = dict(zip(df_offset['ModuleCode'], df_offset['Offset']))
except:
    offset_dict = {}
    
try:
    df_settings = sanitize_df(pd.read_excel(xls, sheet_name="Settings"))
    settings = dict(zip(df_settings['Parameter'], df_settings['Value']))
    MAX_CLASSES_PER_DAY = int(settings.get('MaxClassesPerDay', 3))
    MAX_STAFF_CLASSES_PER_DAY = int(settings.get('MaxStaffClassesPerDay', 4))
except:
    MAX_CLASSES_PER_DAY = 3
    MAX_STAFF_CLASSES_PER_DAY = 4

# Map explicit blanks back to "All" for Module Staff assignments
if 'Term' not in df_mod_staff.columns: 
    df_mod_staff['Term'] = "All"
df_mod_staff['Term'] = df_mod_staff['Term'].replace({'': 'All'})

if 'NumberOfRooms' not in df_rooms.columns: df_rooms['NumberOfRooms'] = 0
df_rooms['NumberOfRooms'] = pd.to_numeric(df_rooms['NumberOfRooms'], errors='coerce').fillna(0)

days_list = df_timeslots['Day'].unique().tolist()
T = [f"{row['Day']}_{row['TimeSlot']}" for idx, row in df_timeslots.iterrows()]
ordered_timeslots = []
for idx, row in df_timeslots.iterrows():
    if row['TimeSlot'] not in ordered_timeslots: ordered_timeslots.append(row['TimeSlot'])
    
all_staff = df_mod_staff['StaffInitials'].unique().tolist()
all_progs = df_mod_prog['ProgrammeCode'].unique().tolist()

# ==========================================
# 4. GLOBAL GROUP SIZING & BASE DEPENDENCIES
# ==========================================
print("[*] Slicing Cohorts & Building Global Dependencies...", flush=True)

def get_target_cohorts(p, mod_code, current_term, ordered_terms, df_mod_prog, offset):
    mod_prog_subset = df_mod_prog[(df_mod_prog['ModuleCode'] == mod_code) & (df_mod_prog['ProgrammeCode'] == p)]
    terms_offered = mod_prog_subset['Term'].tolist()
    
    target_cohorts = [f"{t}Cohort" for t in ordered_terms]
    
    if len(terms_offered) > 1:
        target_cohorts = []
        curr_idx = ordered_terms.index(current_term)
        mapped_idx = (curr_idx - offset) % len(ordered_terms)
        target_cohorts.append(f"{ordered_terms[mapped_idx]}Cohort")
        
        if len(terms_offered) < len(ordered_terms):
            sorted_offered = sorted(terms_offered, key=lambda x: ordered_terms.index(x))
            if current_term == sorted_offered[0]:
                missing_terms = [t for t in ordered_terms if t not in terms_offered]
                for mt in missing_terms:
                    mt_idx = ordered_terms.index(mt)
                    mt_mapped = (mt_idx - offset) % len(ordered_terms)
                    target_cohorts.append(f"{ordered_terms[mt_mapped]}Cohort")
                    
    return list(set(target_cohorts))

prog_cohort_groups = {}
all_groups = set()
group_sizes_for_export = []

prog_cohort_min_cap = {}
for (mod_code, term), group in df_mod_prog.groupby(['ModuleCode', 'Term']):
    progs = group['ProgrammeCode'].tolist()
    offset = int(offset_dict.get(mod_code, 0))
    base_capacity = int(df_rooms.loc[df_modules.loc[mod_code, "RoomType"], "Capacity"])
    if base_capacity <= 0: base_capacity = 9999 
    
    for p in progs:
        target_cohorts = get_target_cohorts(p, mod_code, term, ordered_terms, df_mod_prog, offset)
        for cohort_col in target_cohorts:
            key = (p, cohort_col)
            if key not in prog_cohort_min_cap:
                prog_cohort_min_cap[key] = set()
            prog_cohort_min_cap[key].add(base_capacity)

for (p, cohort_col), caps in prog_cohort_min_cap.items():
    if p not in df_prog.index: continue
    raw_size = df_prog.loc[p, cohort_col]
    if pd.isna(raw_size): continue
    c_size = int(raw_size)
    if c_size <= 0: continue
    
    hcf_val = reduce(math.gcd, list(caps)) if caps else 20
    chunk_size = hcf_val if hcf_val >= 5 else 5 
    
    num_full = c_size // chunk_size
    leftover = c_size % chunk_size
    term_name = cohort_col.replace('Cohort', '')
    
    prog_cohort_groups[(p, cohort_col)] = []
    g_idx = 1
    
    for i in range(num_full):
        g_name = f"{p}-{term_name}-G{g_idx}"
        prog_cohort_groups[(p, cohort_col)].append({'name': g_name, 'size': chunk_size})
        all_groups.add(g_name)
        group_sizes_for_export.append({"Programme": p, "Term Cohort": term_name, "Group Name": g_name, "Size": chunk_size})
        g_idx += 1
        
    if leftover > 0:
        g_name = f"{p}-{term_name}-G{g_idx}"
        prog_cohort_groups[(p, cohort_col)].append({'name': g_name, 'size': leftover})
        all_groups.add(g_name)
        group_sizes_for_export.append({"Programme": p, "Term Cohort": term_name, "Group Name": g_name, "Size": leftover})

module_term_info = []

for (mod_code, term), group in df_mod_prog.groupby(['ModuleCode', 'Term']):
    progs_taking_module = group['ProgrammeCode'].tolist()
    offset = int(offset_dict.get(mod_code, 0))
    
    module_groups = []
    for p in progs_taking_module:
        target_cohorts = get_target_cohorts(p, mod_code, term, ordered_terms, df_mod_prog, offset)
        for cohort_col in target_cohorts:
            if (p, cohort_col) in prog_cohort_groups:
                module_groups.extend(prog_cohort_groups[(p, cohort_col)])
            
    total_students = sum(g['size'] for g in module_groups)
    base_room_type = df_modules.loc[mod_code, "RoomType"]
    base_capacity = int(df_rooms.loc[base_room_type, "Capacity"])
    
    num_sections = math.ceil(total_students / base_capacity) if base_capacity > 0 else 1
    if num_sections == 0: continue
    
    module_term_info.append({
        'mod_code': mod_code,
        'term': term,
        'num_sections': num_sections,
        'progs_taking_module': progs_taking_module,
        'module_groups': module_groups,
        'base_capacity': base_capacity
    })

module_term_info.sort(key=lambda x: x['num_sections'], reverse=True)

global_group_idx = {}
sorted_all_groups = sorted(list(all_groups))
for idx, g_name in enumerate(sorted_all_groups):
    global_group_idx[g_name] = idx

# ==========================================
# 5. THE TERM-ISOLATED OPTIMIZATION LOOP
# ==========================================
global_perfect = True
term_data = {term: {t: {d: [] for d in days_list} for t in ordered_timeslots} for term in ordered_terms}
staff_data = {}
group_data = {}
module_data = {} 
programme_data = {} 

for current_term in ordered_terms:
    try:
        print(f"\n{'='*60}", flush=True)
        print(f"[*] INITIALIZING ENGINE FOR: {current_term.upper()} TERM", flush=True)
        print(f"{'='*60}", flush=True)
        
        term_modules = [info for info in module_term_info if info['term'] == current_term]
        
        if not term_modules:
            print(f"[*] No modules scheduled for {current_term} term. Skipping.", flush=True)
            continue
            
        print(f"[*] PRIORITY QUEUE ({current_term.upper()})", flush=True)
        for info in term_modules:
            gravity_weight = 1000 * (10 ** info['num_sections'])
            print(f"    - {info['mod_code']}: {info['num_sections']} sections  ->  Gravity Penalty: {gravity_weight:,}", flush=True)
        
        # ----------------------------------------------------
        # A. Strict Contiguous Clump Packing
        # ----------------------------------------------------
        S = []
        module_of_section = {}
        staff_of_section = {}
        prog_of_section = {}
        groups_of_section = {}
        valid_rooms_for_section = {}
        section_weight = {}

        for info in term_modules:
            mod_code = info['mod_code']
            num_sections = info['num_sections']
            module_groups = info['module_groups']
            base_capacity = info['base_capacity']
            progs_taking_module = info['progs_taking_module']

            if not module_groups or num_sections == 0:
                continue

            module_groups.sort(key=lambda x: x['name'])
            
            section_names = [f"{mod_code}-{current_term}-{chr(65+i)}" for i in range(num_sections)]
            section_loads = {s: 0 for s in section_names}
            assigned_groups = {s: [] for s in section_names}
            
            num_groups = len(module_groups)
            base_count = num_groups // num_sections if num_sections > 0 else 0
            remainder = num_groups % num_sections if num_sections > 0 else 0
            
            group_idx = 0
            for i, s_name in enumerate(section_names):
                count_for_this_sec = base_count + (1 if i < remainder else 0)
                for _ in range(count_for_this_sec):
                    if group_idx < num_groups:
                        g = module_groups[group_idx]
                        assigned_groups[s_name].append(g['name'])
                        section_loads[s_name] += g['size']
                        group_idx += 1
            
            mod_staff_df = df_mod_staff[(df_mod_staff['ModuleCode'] == mod_code) & 
                                        (df_mod_staff['Term'].isin(['All', current_term]))]
            staff_all = mod_staff_df[mod_staff_df['AllocationType'].astype(str).str.title() == 'All']['StaffInitials'].tolist()
            staff_split_df = mod_staff_df[mod_staff_df['AllocationType'].astype(str).str.title() == 'Split']
            
            split_quotas = {}
            if num_sections > 1 and not staff_split_df.empty:
                allocated = 0
                fallback_pct = 100.0 / len(staff_split_df)
                for _, row in staff_split_df.iterrows():
                    st = row['StaffInitials']
                    percent = pd.to_numeric(row['SplitPercentage'], errors='coerce')
                    if pd.isna(percent):
                        percent = fallback_pct
                    q = int(round(num_sections * (percent / 100.0)))
                    split_quotas[st] = q
                    allocated += q
                
                # Assign remainder safely to the first staff member if math is imperfect
                diff = num_sections - allocated
                if diff != 0:
                    first_staff = staff_split_df.iloc[0]['StaffInitials']
                    split_quotas[first_staff] += diff

            for s in section_names:
                S.append(s)
                section_weight[s] = num_sections 
                module_of_section[s] = mod_code
                prog_of_section[s] = progs_taking_module 
                groups_of_section[s] = assigned_groups[s]
                
                actual_packed_size = section_loads[s]
                required_cap = max(base_capacity, actual_packed_size)
                valid_rooms = [r for r in df_rooms.index if df_rooms.loc[r, "Capacity"] >= required_cap]
                
                if not valid_rooms:
                    largest_room = df_rooms['Capacity'].idxmax()
                    valid_rooms = [largest_room]
                    
                valid_rooms_for_section[s] = valid_rooms
                
                assigned_staff = staff_all.copy() 
                if num_sections == 1:
                    assigned_staff.extend(staff_split_df['StaffInitials'].tolist())
                else:
                    for st in split_quotas:
                        if split_quotas[st] > 0:
                            assigned_staff.append(st)
                            split_quotas[st] -= 1
                            break 
                            
                staff_of_section[s] = assigned_staff

        # ----------------------------------------------------
        # B. Term-Isolated Mathematical Optimization
        # ----------------------------------------------------
        print(f"\n[*] Generating isolated constraints for {current_term.upper()}...", flush=True)
        model = LpProblem(f"Timetable_{current_term}", LpMinimize)

        x = LpVariable.dicts(f"Schedule_{current_term}", [(s, t, r) for s in S for t in T for r in valid_rooms_for_section[s]], cat='Binary')
        slack_drop = LpVariable.dicts(f"Drop_{current_term}", S, cat='Binary')
        
        term_groups = set()
        term_staff = set()
        for s in S:
            term_groups.update(groups_of_section[s])
            term_staff.update(staff_of_section[s])
            
        slack_over = LpVariable.dicts(f"GrpOver_{current_term}", [(g, d) for g in term_groups for d in days_list], lowBound=0, cat='Integer')
        slack_staff_over = LpVariable.dicts(f"StaffOver_{current_term}", [(st, d) for st in term_staff for d in days_list], lowBound=0, cat='Integer')

        wasted_seat_penalty = lpSum([
            x[(s, t, r)] * (df_rooms.loc[r, "Capacity"] - df_rooms.loc[df_modules.loc[module_of_section[s], "RoomType"], "Capacity"])
            for s in S for t in T for r in valid_rooms_for_section[s]
        ])

        alignment_penalty = lpSum([
            x[(s, t, r)] * 5 * (1 if ordered_timeslots.index(t.split('_')[1]) != (ord(s[-1]) - 65) % len(ordered_timeslots) else 0)
            for s in S for t in T for r in valid_rooms_for_section[s]
        ])

        model += (
            lpSum([slack_drop[s] * float(1000 * (10 ** section_weight[s])) for s in S]) + 
            lpSum([slack_over[(g, d)] * 1000 for g in term_groups for d in days_list]) +
            lpSum([slack_staff_over[(st, d)] * 1000 for st in term_staff for d in days_list]) +
            wasted_seat_penalty +
            alignment_penalty
        ), "Objective"

        for s in S:
            model += lpSum([x[(s, t, r)] for t in T for r in valid_rooms_for_section[s]]) + slack_drop[s] == 1 

        for t in T:
            for staff in term_staff:
                staff_secs = [s for s in S if staff in staff_of_section[s]]
                if len(staff_secs) > 1: 
                    model += lpSum([x[(s, t, r)] for s in staff_secs for r in valid_rooms_for_section[s]]) <= 1 

            for g in term_groups:
                grp_secs = [s for s in S if g in groups_of_section[s]]
                if len(grp_secs) > 1: 
                    model += lpSum([x[(s, t, r)] for s in grp_secs for r in valid_rooms_for_section[s]]) <= 1 

        for g in term_groups:
            for d in days_list:
                day_ts = [t for t in T if t.startswith(d)]
                grp_secs = [s for s in S if g in groups_of_section[s]]
                if grp_secs and day_ts:
                    model += lpSum([x[(s, t, r)] for s in grp_secs for t in day_ts for r in valid_rooms_for_section[s]]) <= MAX_CLASSES_PER_DAY + slack_over[(g, d)]

        for st in term_staff:
            for d in days_list:
                day_ts = [t for t in T if t.startswith(d)]
                staff_secs = [s for s in S if st in staff_of_section[s]]
                if staff_secs and day_ts:
                    model += lpSum([x[(s, t, r)] for s in staff_secs for t in day_ts for r in valid_rooms_for_section[s]]) <= MAX_STAFF_CLASSES_PER_DAY + slack_staff_over[(st, d)]

        for s in S:
            for t in T:
                d = t.split('_')[0]
                blocked = False
                for staff in staff_of_section[s]:
                    if staff in restricted_staff_set and (staff, current_term, t) not in permitted_staff: 
                        blocked = True
                for p in prog_of_section[s]:
                    if p in restricted_prog_set and (p, d) not in permitted_progs: 
                        blocked = True
                if blocked:
                    model += lpSum([x[(s, t, r)] for r in valid_rooms_for_section[s]]) == 0

        for r in df_rooms.index:
            max_rooms_of_type = df_rooms.loc[r, "NumberOfRooms"]
            if max_rooms_of_type > 0: 
                for t in T:
                    secs_using_this_room = [s for s in S if r in valid_rooms_for_section[s]]
                    if len(secs_using_this_room) > 1:
                        model += lpSum([x[(s, t, r)] for s in secs_using_this_room]) <= max_rooms_of_type

        # ----------------------------------------------------
        # C. Solving the Isolated Term
        # ----------------------------------------------------
        print(f"[*] Solving {current_term.upper()} model... (Time Limit: 15 mins, Gap: 5%)", flush=True)
        try:
            model.solve(PULP_CBC_CMD(gapRel=0.05, timeLimit=900, msg=True))
        except TypeError:
            try:
                model.solve(PULP_CBC_CMD(fracGap=0.05, maxSeconds=900, msg=True))
            except TypeError:
                model.solve(PULP_CBC_CMD(msg=True, options=['ratio', '0.05', 'sec', '900']))

        print(f"\n--- DIAGNOSTICS FOR {current_term.upper()} ---", flush=True)
        term_perfect = True
        for s in S:
            if slack_drop[s].varValue == 1:
                print(f"[!] DROPPED: {s} (Mathematical impossibility based on limits)", flush=True)
                term_perfect = False
        for g in term_groups:
            for d in days_list:
                if slack_over[(g, d)].varValue and slack_over[(g, d)].varValue > 0:
                    print(f"[!] GROUP OVERLOAD: {g} exceeds limits on {d}", flush=True)
                    term_perfect = False
        for st in term_staff:
            for d in days_list:
                over_val = slack_staff_over[(st, d)].varValue
                if over_val and over_val > 0:
                    print(f"[!] STAFF OVERLOAD: {st} exceeds limits on {d}", flush=True)
                    term_perfect = False

        if term_perfect: 
            print(f"[+] SUCCESS! {current_term.upper()} Constraints met flawlessly.", flush=True)
        else:
            global_perfect = False
            
        # ----------------------------------------------------
        # D. Parse and Append to Global Excel Dictionaries
        # ----------------------------------------------------
        for s in S:
            if slack_drop[s].varValue != 1:
                for t_slot in T:
                    for r in valid_rooms_for_section[s]:
                        if x[(s, t_slot, r)].varValue == 1:
                            day_part = t_slot.split('_')[0]
                            time_part = t_slot.split('_')[1]
                            mod_code = module_of_section[s]
                            mod_name = df_modules.loc[mod_code, "ModuleName"]
                            staff_list = staff_of_section[s]
                            
                            staff_names = ", ".join([df_staff.loc[st, "FullName"] for st in staff_list])
                            groups_str = ", ".join(groups_of_section[s])
                            
                            base_req_room = df_modules.loc[mod_code, "RoomType"]
                            room_note = f"{r}" if r == base_req_room else f"{r} (from {base_req_room})"
                            
                            class_info = {
                                "mod_code": mod_code,
                                "mod_name": mod_name,
                                "staff_names": staff_names,
                                "section": s,
                                "groups": groups_str,
                                "room": room_note,
                                "day": day_part,
                                "time": time_part,
                                "term": current_term
                            }
                            
                            term_data[current_term][time_part][day_part].append(class_info)
                            
                            for p in prog_of_section[s]:
                                if p not in programme_data:
                                    programme_data[p] = {tm: {tt: {dd: [] for dd in days_list} for tt in ordered_timeslots} for tm in ordered_terms}
                                programme_data[p][current_term][time_part][day_part].append(class_info)
                            
                            for st in staff_list:
                                if st not in staff_data:
                                    staff_data[st] = {tm: {tt: {dd: [] for dd in days_list} for tt in ordered_timeslots} for tm in ordered_terms}
                                staff_data[st][current_term][time_part][day_part].append(class_info)
                                
                            for g in groups_of_section[s]:
                                if g not in group_data:
                                    group_data[g] = {tm: {tt: {dd: [] for dd in days_list} for tt in ordered_timeslots} for tm in ordered_terms}
                                group_data[g][current_term][time_part][day_part].append(class_info)
                                
                            if mod_code not in module_data:
                                module_data[mod_code] = []
                            module_data[mod_code].append(class_info)

    except Exception as e:
        print(f"\n[!] FATAL ERROR DURING MATH OPTIMIZATION FOR {current_term.upper()}:\n{e}", flush=True)
        traceback.print_exc()

print(f"\n{'='*60}", flush=True)
if global_perfect:
    print("[+] ALL TERMS PROCESSED PERFECTLY.", flush=True)
else:
    print("[-] SOLVER FINISHED WITH SOME COMPROMISES (Check above for dropped classes)", flush=True)
print(f"{'='*60}\n", flush=True)

# ==========================================
# 6. EXCEL DASHBOARD GENERATION
# ==========================================
print("[*] Generating formatted Excel grids...", flush=True)

out_dir = os.path.dirname(file_path)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_filename = f"Timetable_Draft_{timestamp}.xlsx"
out_path = os.path.join(out_dir, out_filename)

wb = Workbook()
wb.remove(wb.active)

bold_font = InlineFont(b=True)
reg_font = InlineFont(b=False)

def write_grid(ws, title, grid, start_row):
    ws.cell(row=start_row, column=1, value=title).font = Font(size=14, bold=True)
    start_row += 2
    
    ws.cell(row=start_row, column=1, value="Time / Day").font = Font(bold=True)
    for col_idx, day in enumerate(days_list, start=2):
        ws.cell(row=start_row, column=col_idx, value=day).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = 35
        
    start_row += 1
    
    for time_idx, time_part in enumerate(ordered_timeslots):
        row_idx = start_row + time_idx
        ws.cell(row=row_idx, column=1, value=time_part).font = Font(bold=True)
        ws.column_dimensions['A'].width = 20
        
        max_lines_in_row = 1
        for col_idx, day in enumerate(days_list, start=2):
            classes = grid[time_part][day]
            if not classes: continue
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            rt_elements = []
            for idx, c in enumerate(classes):
                line1 = f"{c['mod_code']} - {c['mod_name']}\n"
                line2 = f"{c['staff_names']}\n{c['section']}\n{c['groups']}\n{c['room']}"
                if idx < len(classes) - 1:
                    line2 += "\n\n"
                    
                rt_elements.append(TextBlock(bold_font, line1))
                rt_elements.append(TextBlock(reg_font, line2))
                
            try:
                cell.value = CellRichText(*rt_elements)
            except:
                text_lines = []
                for c in classes:
                    text_lines.append(f"{c['mod_code']} - {c['mod_name']}\n{c['staff_names']}\n{c['section']}\n{c['groups']}\n{c['room']}")
                cell.value = "\n\n".join(text_lines)
                
            lines_in_cell = sum(6 for _ in classes) + (len(classes) - 1)
            if lines_in_cell > max_lines_in_row:
                max_lines_in_row = lines_in_cell
                
        ws.row_dimensions[row_idx].height = max_lines_in_row * 14.5

    return start_row + len(ordered_timeslots) + 2

def write_linear_column(ws, title, class_list):
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    ws.column_dimensions['A'].width = 60
    
    def sort_key(c):
        t_idx = ordered_terms.index(c['term']) if c['term'] in ordered_terms else 99
        d_idx = days_list.index(c['day']) if c['day'] in days_list else 99
        tm_idx = ordered_timeslots.index(c['time']) if c['time'] in ordered_timeslots else 99
        return (t_idx, d_idx, tm_idx, c['section'])
        
    sorted_classes = sorted(class_list, key=sort_key)
    
    current_row = 3
    for c in sorted_classes:
        line1 = f"{c['term']} Term | {c['day']} | {c['time']}\n"
        line2 = f"Section: {c['section']}\nStaff: {c['staff_names']}\nGroups: {c['groups']}\nRoom: {c['room']}"
        
        rt_elements = [TextBlock(bold_font, line1), TextBlock(reg_font, line2)]
        
        cell = ws.cell(row=current_row, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        try:
            cell.value = CellRichText(*rt_elements)
        except:
            cell.value = line1 + line2
            
        ws.row_dimensions[current_row].height = 5 * 14.5
        current_row += 2

# 1. Groups Summary Sheet
ws_groups = wb.create_sheet(title="Groups", index=0)
ws_groups.append(["Programme", "Term Cohort", "Group Name", "Number of Students"])
for cell in ws_groups[1]:
    cell.font = Font(bold=True)

group_sizes_for_export = sorted(group_sizes_for_export, key=lambda x: (x["Programme"], x["Term Cohort"], x["Group Name"]))
for r in group_sizes_for_export:
    ws_groups.append([r["Programme"], r["Term Cohort"], r["Group Name"], r["Size"]])

ws_groups.column_dimensions['A'].width = 20
ws_groups.column_dimensions['B'].width = 15
ws_groups.column_dimensions['C'].width = 25
ws_groups.column_dimensions['D'].width = 20

# 2. Term Sheets
for term in ordered_terms:
    if any(term_data[term][tm][d] for tm in ordered_timeslots for d in days_list):
        ws = wb.create_sheet(title=f"Term - {term}")
        write_grid(ws, f"{term} Term Timetable", term_data[term], 1)

# 3. Programme Sheets (MATRIX LAYOUT)
sorted_progs = sorted(programme_data.keys())
for p in sorted_progs:
    ws = wb.create_sheet(title=f"Prog - {p}"[:31])
    current_row = 1
    prog_name_full = df_prog.loc[p, "ProgrammeName"]
    ws.cell(row=current_row, column=1, value=f"Timetable for {p} - {prog_name_full}").font = Font(size=16, bold=True)
    current_row += 2
    for term in ordered_terms:
        if any(programme_data[p][term][tm][d] for tm in ordered_timeslots for d in days_list):
            current_row = write_grid(ws, f"{term} Term", programme_data[p][term], current_row)

# 4. Module Sheets (LINEAR LAYOUT)
sorted_modules = sorted(module_data.keys())
for m in sorted_modules:
    ws = wb.create_sheet(title=f"Mod - {m}"[:31]) 
    mod_name_full = df_modules.loc[m, "ModuleName"]
    write_linear_column(ws, f"Timetable for {m} - {mod_name_full}", module_data[m])

# 5. Group Sheets
sorted_groups = sorted(group_data.keys())
for g in sorted_groups:
    ws = wb.create_sheet(title=f"{g}"[:31])
    current_row = 1
    for term in ordered_terms:
        if any(group_data[g][term][tm][d] for tm in ordered_timeslots for d in days_list):
            current_row = write_grid(ws, f"{term} Term", group_data[g][term], current_row)

# 6. Staff Sheets
sorted_staff = sorted(staff_data.keys())
for st in sorted_staff:
    ws = wb.create_sheet(title=f"{st}"[:31])
    current_row = 1
    st_name = df_staff.loc[st, "FullName"]
    ws.cell(row=current_row, column=1, value=f"Timetable for {st_name}").font = Font(size=16, bold=True)
    current_row += 2
    for term in ordered_terms:
        if any(staff_data[st][term][tm][d] for tm in ordered_timeslots for d in days_list):
            current_row = write_grid(ws, f"{term} Term", staff_data[st][term], current_row)

wb.save(out_path)
print(f"[*] Done! Formatted Timetable successfully saved to:\n    {out_path}", flush=True)